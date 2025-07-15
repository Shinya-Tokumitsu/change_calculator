import os
import re
import sys
import csv
import pandas as pd
import numpy  as np
import matplotlib.pyplot as plt
import math
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from dataclasses import dataclass
from typing import List
from abc import ABC, abstractmethod
from read_excel import *
from file_utils import *
from least_squares import *

@dataclass
class BaseResult(InputData):
  """
  InputDataクラスを継承し計算結果を保持するクラス.
  すべての計算結果に共通する属性と振る舞いを定義する「設計図」．
  """
  std_situation : str
  std_coord: pd.DataFrame
  change_radius: List[float]
  lsm_change_radius: List[float]
  change_diameter: List[float]
  rad_degrees: List[float]
  dia_degrees: List[float]
  std_rod_dist : float
  lsm_cx : float
  lsm_cy : float
  lsm_r  : float
  
  @property
  @abstractmethod
  def effective_change_diameter(self) -> List[float]:
      """計算に用いる内径変化量データを返す（生 or FFT）"""
      pass

  @property
  @abstractmethod
  def effective_lsm_change_radius(self) -> List[float]:
      """計算に用いる補正半径変化量データを返す（生 or FFT）"""
      pass
  
  @property
  def roundness(self) -> float:
      """真円度を計算して返す"""   
      return max(self.effective_lsm_change_radius) - min(self.effective_lsm_change_radius)

  @property
  def simple_roundness(self) -> float:
      """簡易真円度を計算して返す"""
      return (max(self.effective_change_diameter) - min(self.effective_change_diameter)) / 2

  @property
  def sliding_distance(self) -> float:
      """滑り量を計算して返す"""   
      non_std_rod_dist = np.sqrt((self.sorted_coord["rod_x"].iloc[0] - self.sorted_coord["rod_x"].iloc[-1])**2
                    + (self.sorted_coord["rod_y"].iloc[0] - self.sorted_coord["rod_y"].iloc[-1])**2)
      sliding_distance = (self.std_rod_dist - non_std_rod_dist) * 10**3
      return sliding_distance
  
  @property
  def amount_of_pull_in(self) -> float:
      """引き込み量を計算して返す"""
      return -min(self.effective_change_diameter)

  @property
  def close_in(self) -> float:
      """クローズインを計算して返す"""
      return max(self.amount_of_pull_in, self.sliding_distance)

  @property
  def max_change_dia(self) -> float:
      """内径変化量の最大値を返す"""
      return max(self.effective_change_diameter)

  @property
  def min_change_dia(self) -> float:
      """内径変化量の最小値を返す"""
      return min(self.effective_change_diameter)
  
  def _merge_header(self, worksheet: Worksheet, text: str, start_col: int, end_col: int):
      """Excelシートのヘッダーを結合・装飾するプライベートメソッド"""
      thin_border_side = Side(style='thin', color='000000')
      box_border = Border(top=thin_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)
      for col in range(start_col, end_col + 1):
          cell = worksheet.cell(row=1, column=col)
          cell.border = box_border
      header_cell = worksheet.cell(row=1, column=start_col)
      header_cell.value = text
      header_cell.alignment = Alignment(horizontal='center', vertical='center')
      header_cell.font = Font(bold=True)
      worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

  @abstractmethod
  def write_to_output_excel_sheet(self, writer: pd.ExcelWriter):
    """自分自身をExcelブックに1シートとして書き出す"""
    sheet_name = self.situation[:30]
    worksheet: Worksheet = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = worksheet
    current_col = 1

    # 座標データブロック (共通)
    self.std_coord.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, startcol=current_col-1)
    std_coord_col_start = current_col
    current_col += self.std_coord.shape[1]

    self.sorted_coord.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, startcol=current_col-1)
    coord_col_start = current_col
    current_col += self.sorted_coord.shape[1] + 1
  
    # 内径変化量ブロック (子クラスでDataFrameを生成)
    dia_df = self._create_dia_df()
    dia_col_start = current_col
    dia_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, startcol=current_col-1)
    current_col += dia_df.shape[1] + 1

    # 半径変化量ブロック (子クラスでDataFrameを生成)
    rad_df = self._create_rad_df()
    rad_col_start = current_col
    rad_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, startcol=current_col-1)
    current_col += rad_df.shape[1] + 1

    # 結果ブロック (共通)
    results_data = [
        ("クローズイン", self.close_in), ("滑り量", self.sliding_distance),
        ("引き込み量", self.amount_of_pull_in), ("簡易真円度", self.simple_roundness),
        ("内径変化量の最大値", self.max_change_dia), ("内径変化量の最小値", self.min_change_dia),
        ("真円度", self.roundness), ("最小二乗円 Cx", self.lsm_cx),
        ("最小二乗円 Cy", self.lsm_cy), ("最小二乗円 半径", self.lsm_r)
    ]
    results_df = pd.DataFrame(results_data, columns=["result", "value"])
    results_col_start = current_col
    results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=1, startcol=current_col-1)
    
    # ヘッダー結合 (共通)
    self._merge_header(worksheet, self.std_situation, std_coord_col_start, std_coord_col_start + 3)
    self._merge_header(worksheet, self.situation, coord_col_start, coord_col_start + 3)
    self._merge_header(worksheet, "内径変化量", dia_col_start, dia_col_start + dia_df.shape[1] - 1)
    self._merge_header(worksheet, "半径変化量", rad_col_start, rad_col_start + rad_df.shape[1] - 1)
    self._merge_header(worksheet, "結果", results_col_start, results_col_start + results_df.shape[1] - 1)
    

@dataclass
class NonFFTResult(BaseResult):
  """FFTがOFFの場合の結果クラス"""
  @property
  def effective_change_diameter(self) -> List[float]:
    return self.change_diameter
  
  @property
  def effective_lsm_change_radius(self) -> List[float]:
    return self.lsm_change_radius
  
  def _create_dia_df(self) -> pd.DataFrame:
    return pd.DataFrame({
        "角度[°]" : self.dia_degrees,
        "内径変化量[㎛]": self.change_diameter
    })

  def _create_rad_df(self) -> pd.DataFrame:
    return pd.DataFrame({
        "角度[°]": self.rad_degrees,
        "半径変化量[㎛]": self.change_radius,
        "半径変化量（補正）[㎛]": self.lsm_change_radius
    })


@dataclass
class FFTResult(BaseResult):
  """FFTがONの場合の結果クラス"""
  fft_change_radius: List[float]
  fft_lsm_change_radius: List[float]
  fft_change_diameter: List[float]

  @property
  def effective_change_diameter(self) -> List[float]:
    return self.fft_change_diameter
  
  @property
  def effective_lsm_change_radius(self) -> List[float]:
    return self.fft_lsm_change_radius
  
  def _create_dia_df(self) -> pd.DataFrame:
    return pd.DataFrame({
        "角度[°]" : self.dia_degrees,
        "内径変化量[㎛]": self.change_diameter,
        "FFT内径変化量[㎛]": self.fft_change_diameter
    })

  def _create_rad_df(self) -> pd.DataFrame:
    return pd.DataFrame({
        "角度[°]": self.rad_degrees,
        "半径変化量[㎛]": self.change_radius,
        "FFT半径変化量[㎛]": self.fft_change_radius,
        "半径変化量（補正）[㎛]": self.lsm_change_radius,
        "FFT半径変化量（補正）[㎛]": self.fft_lsm_change_radius
    })
